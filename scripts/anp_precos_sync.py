#!/usr/bin/env python3
"""
RotaPosto – ANP Precos Sync
============================
Baixa a planilha semanal de preços da ANP e faz upload para o
Cloudflare KV, particionado por UF para respeitar o limite de 25 MB.

Chaves no KV:
  precos:municipios          → JSON com médias por município (todos UFs)
  precos:postos:SP           → JSON com preços por CNPJ em SP
  precos:postos:RJ           → JSON com preços por CNPJ no RJ
  ... (uma chave por UF)
  precos:meta                → metadados: semana, total, timestamp

Uso:
  python3 scripts/anp_precos_sync.py

Env vars necessárias:
  CLOUDFLARE_API_TOKEN       → token CF com permissão KV Write
  CLOUDFLARE_ACCOUNT_ID      → ID da conta CF
  CLOUDFLARE_KV_NAMESPACE_ID → ID do namespace KV
"""

import os
import sys
import json
import gzip
import hashlib
import urllib.request
import urllib.error
from datetime import datetime, timedelta
from collections import defaultdict

# ─────────────────────────────────────────────────────────────────────────────
# Dependências opcionais
# ─────────────────────────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# Configuração
# ─────────────────────────────────────────────────────────────────────────────
CF_TOKEN     = os.environ.get("CLOUDFLARE_API_TOKEN", "")
CF_ACCOUNT   = os.environ.get("CLOUDFLARE_ACCOUNT_ID", "")
CF_KV_NS     = os.environ.get("CLOUDFLARE_KV_NAMESPACE_ID", "")
ANP_BASE_URL = "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos/arquivos-lpc"

# ─────────────────────────────────────────────────────────────────────────────
# Mapeamentos
# ─────────────────────────────────────────────────────────────────────────────
PRODUTO_MAP = {
    "GASOLINA COMUM":     "gasolina",
    "GASOLINA ADITIVADA": "gasolinaAditivada",
    "ETANOL":             "etanol",
    "DIESEL S10":         "dieselS10",
    "DIESEL S500":        "diesel",
    "GNV":                "gnv",
    "GLP":                "glp",
}

UF_SIGLA = {
    "ACRE":"AC","ALAGOAS":"AL","AMAPA":"AP","AMAZONAS":"AM","BAHIA":"BA",
    "CEARA":"CE","DISTRITO FEDERAL":"DF","ESPIRITO SANTO":"ES","GOIAS":"GO",
    "MARANHAO":"MA","MATO GROSSO":"MT","MATO GROSSO DO SUL":"MS",
    "MINAS GERAIS":"MG","PARA":"PA","PARAIBA":"PB","PARANA":"PR",
    "PERNAMBUCO":"PE","PIAUI":"PI","RIO DE JANEIRO":"RJ",
    "RIO GRANDE DO NORTE":"RN","RIO GRANDE DO SUL":"RS","RONDONIA":"RO",
    "RORAIMA":"RR","SANTA CATARINA":"SC","SAO PAULO":"SP",
    "SERGIPE":"SE","TOCANTINS":"TO"
}


# ─────────────────────────────────────────────────────────────────────────────
# Funções utilitárias
# ─────────────────────────────────────────────────────────────────────────────
def gerar_urls_anp(semanas_atras: int = 3) -> list[tuple[str, str, str]]:
    """
    Gera URLs das últimas N semanas da ANP.
    A ANP usa ciclo domingo→sábado:
      Ex: 21/06/2026 (dom) a 27/06/2026 (sab)
    """
    urls = []
    hoje = datetime.now()
    # weekday(): 0=seg, 1=ter, ..., 5=sab, 6=dom
    dia_semana = hoje.weekday()
    # Dias desde o último domingo (domingo = weekday 6 → 0 dias; segunda = 1 dia; sab = 6 dias)
    dias_desde_domingo = (dia_semana + 1) % 7
    ultimo_domingo = hoje - timedelta(days=dias_desde_domingo)

    for i in range(semanas_atras):
        domingo = ultimo_domingo - timedelta(weeks=i)
        sabado  = domingo + timedelta(days=6)
        ano = sabado.year
        url = (
            f"{ANP_BASE_URL}/{ano}/"
            f"revendas_lpc_{domingo.strftime('%Y-%m-%d')}_{sabado.strftime('%Y-%m-%d')}.xlsx"
        )
        urls.append((domingo.strftime("%Y-%m-%d"), sabado.strftime("%Y-%m-%d"), url))

    return urls


def baixar_xlsx(url: str) -> bytes | None:
    """Baixa o arquivo xlsx da URL, retorna bytes ou None."""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
        "Accept-Language": "pt-BR,pt;q=0.9",
        "Referer": "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos"
    }
    req = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            if resp.status == 200:
                data = resp.read()
                print(f"  ✓ Baixado {len(data)/1024:.0f} KB de {url.split('/')[-1]}")
                return data
    except urllib.error.HTTPError as e:
        print(f"  ✗ HTTP {e.code} em {url.split('/')[-1]}")
    except Exception as e:
        print(f"  ✗ Erro: {e}")
    return None


def parsear_xlsx(data: bytes) -> tuple[dict, dict, str]:
    """
    Parseia o XLSX da ANP.
    Retorna: (por_cnpj, por_municipio, semana_str)
    
    por_cnpj[cnpj] = {n, m, u, b, p:{gasolina:X,...}, dt}
    por_municipio[uf:municipio] = {gasolina:X,...}  ← médias
    """
    import io
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    headers = list(next(ws.iter_rows(min_row=10, max_row=10, values_only=True)))

    por_cnpj: dict[str, dict] = {}
    por_municipio: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))
    semana = ""

    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
        d = dict(zip(headers, row))
        if not d.get("CNPJ") or not d.get("PREÇO DE REVENDA"):
            continue

        try:
            cnpj = str(int(d["CNPJ"])).zfill(14)
            preco = float(d["PREÇO DE REVENDA"])
        except (ValueError, TypeError):
            continue

        produto_raw = str(d.get("PRODUTO") or "").upper()
        produto = PRODUTO_MAP.get(produto_raw, "")
        municipio = str(d.get("MUNICÍPIO") or "").upper()
        estado_full = str(d.get("ESTADO") or "").upper()
        uf = UF_SIGLA.get(estado_full, estado_full[:2])

        if not produto or preco <= 0:
            continue

        data_coleta = d.get("DATA DA COLETA")
        if not semana and data_coleta:
            semana = str(data_coleta)[:10]

        # Por posto (CNPJ)
        if cnpj not in por_cnpj:
            fantasia = str(d.get("FANTASIA") or "")
            razao    = str(d.get("RAZÃO") or "")
            nome = (fantasia if fantasia and fantasia not in ("None", "") else razao)[:40]
            por_cnpj[cnpj] = {
                "n":  nome,
                "m":  municipio,
                "u":  uf,
                "b":  str(d.get("BANDEIRA") or ""),
                "p":  {},
                "dt": str(data_coleta)[:10] if data_coleta else "",
            }
        por_cnpj[cnpj]["p"][produto] = preco

        # Por município (para calcular médias)
        chave = f"{uf}:{municipio}"
        por_municipio[chave][produto].append(preco)

    wb.close()

    # Calcular médias dos municípios
    municipios_media: dict[str, dict] = {}
    for chave, produtos in por_municipio.items():
        municipios_media[chave] = {
            prod: round(sum(precos) / len(precos), 2)
            for prod, precos in produtos.items()
        }

    return por_cnpj, municipios_media, semana


# ─────────────────────────────────────────────────────────────────────────────
# Cloudflare KV API
# ─────────────────────────────────────────────────────────────────────────────
def kv_put(key: str, value: str | bytes, expiration_ttl: int | None = None) -> bool:
    """
    Grava um valor no KV via API REST da Cloudflare.
    Usa multipart/form-data para suportar valores grandes.
    """
    url = (
        f"https://api.cloudflare.com/client/v4/accounts/{CF_ACCOUNT}"
        f"/storage/kv/namespaces/{CF_KV_NS}/values/{urllib.parse.quote(key, safe='')}"
    )

    if isinstance(value, str):
        value = value.encode("utf-8")

    # Gzip para economizar espaço (KV limite: 25 MB)
    compressed = gzip.compress(value, compresslevel=6)
    use_compressed = len(compressed) < len(value)
    body = compressed if use_compressed else value

    params = ""
    if expiration_ttl:
        params = f"?expiration_ttl={expiration_ttl}"

    headers = {
        "Authorization": f"Bearer {CF_TOKEN}",
        "Content-Type": "application/octet-stream",
    }
    if use_compressed:
        headers["Content-Encoding"] = "gzip"

    req = urllib.request.Request(url + params, data=body, headers=headers, method="PUT")
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read())
            if result.get("success"):
                size_kb = len(body) / 1024
                comp_str = f" (gzip {size_kb:.1f} KB)" if use_compressed else f" ({size_kb:.1f} KB)"
                print(f"  ✓ KV PUT '{key}'{comp_str}")
                return True
    except urllib.error.HTTPError as e:
        body_err = e.read().decode()
        print(f"  ✗ KV PUT '{key}' HTTP {e.code}: {body_err[:200]}")
    except Exception as e:
        print(f"  ✗ KV PUT '{key}': {e}")
    return False


# Necessário para urllib.parse.quote
import urllib.parse


def kv_put_json(key: str, obj: dict | list, ttl_days: int = 8) -> bool:
    """Serializa dict/list como JSON compacto e grava no KV."""
    payload = json.dumps(obj, ensure_ascii=False, separators=(",", ":"))
    return kv_put(key, payload, expiration_ttl=ttl_days * 86400)


def kv_get_meta() -> dict | None:
    """Lê os metadados do KV para verificar se já sincronizou esta semana."""
    url = (
        f"https://api.cloudflare.com/client/v4/accounts/{CF_ACCOUNT}"
        f"/storage/kv/namespaces/{CF_KV_NS}/values/precos:meta"
    )
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {CF_TOKEN}"})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read())
    except:
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("RotaPosto – ANP Precos Sync")
    print(f"Hora: {datetime.now().isoformat()}")
    print("=" * 60)

    # Validar credenciais
    if not CF_TOKEN or not CF_ACCOUNT or not CF_KV_NS:
        print("\n⚠️  Credenciais Cloudflare não configuradas!")
        print("  Export: CLOUDFLARE_API_TOKEN, CLOUDFLARE_ACCOUNT_ID, CLOUDFLARE_KV_NAMESPACE_ID")

        # Modo dry-run: só mostra o que faria
        if "--dry-run" in sys.argv or os.environ.get("DRY_RUN"):
            print("\n🔍 Modo DRY-RUN: testando parse sem upload...")
        else:
            sys.exit(1)

    dry_run = "--dry-run" in sys.argv or os.environ.get("DRY_RUN") or not CF_TOKEN

    # Verificar se já sincronizou esta semana
    if not dry_run:
        meta = kv_get_meta()
        if meta:
            semana_atual = meta.get("semana", "")
            hoje = datetime.utcnow().strftime("%Y-%m-%d")
            # Não re-sincronizar na mesma semana (exceto se --force)
            if semana_atual and "--force" not in sys.argv:
                print(f"\n✓ KV já contém dados da semana {semana_atual}. Use --force para re-sincronizar.")
                # Verificar se é semana nova (a última semana difere)
                urls = gerar_urls_anp(1)
                if semana_atual >= urls[0][0]:
                    print("  Dados já estão atualizados!")
                    return 0

    # Tentar baixar o xlsx das últimas 3 semanas
    print("\n1. Buscando planilha ANP semanal...")
    urls = gerar_urls_anp(semanas_atras=3)
    data_xlsx = None
    semana_ini = ""
    semana_fim = ""

    for s_ini, s_fim, url in urls:
        print(f"   Tentando semana {s_ini} → {s_fim}...")
        data_xlsx = baixar_xlsx(url)
        if data_xlsx:
            semana_ini = s_ini
            semana_fim = s_fim
            break

    if not data_xlsx:
        print("\n✗ Não foi possível baixar nenhuma planilha ANP. Abortando.")
        return 1

    # Parsear xlsx
    print("\n2. Parseando planilha...")
    por_cnpj, municipios_media, semana_coleta = parsear_xlsx(data_xlsx)
    del data_xlsx  # liberar memória

    print(f"   ✓ {len(por_cnpj):,} postos com preço")
    print(f"   ✓ {len(municipios_media):,} municípios cobertos")
    print(f"   ✓ Semana de coleta: {semana_coleta}")

    # Organizar por UF
    postos_por_uf: dict[str, dict] = defaultdict(dict)
    for cnpj, posto in por_cnpj.items():
        postos_por_uf[posto["u"]][cnpj] = posto

    print(f"   ✓ {len(postos_por_uf)} UFs com dados")
    for uf, postos in sorted(postos_por_uf.items()):
        print(f"     {uf}: {len(postos):,} postos")

    if dry_run:
        # Mostrar tamanhos
        print("\n[DRY-RUN] Tamanhos dos payloads:")
        mun_json = json.dumps({"s": semana_coleta, "m": municipios_media}, separators=(",",":"))
        print(f"  precos:municipios → {len(mun_json)/1024:.1f} KB")
        for uf, postos in sorted(postos_por_uf.items()):
            j = json.dumps(postos, separators=(",",":"))
            print(f"  precos:postos:{uf} → {len(j)/1024:.1f} KB ({len(postos)} postos)")
        print("\n✓ Dry-run concluído (sem uploads)")
        return 0

    # Upload para Cloudflare KV
    print("\n3. Fazendo upload para Cloudflare KV...")
    erros = 0

    # 3a. Metadados
    meta = {
        "semana": semana_coleta,
        "semanaIni": semana_ini,
        "semanaFim": semana_fim,
        "totalPostos": len(por_cnpj),
        "totalMunicipios": len(municipios_media),
        "ufs": sorted(postos_por_uf.keys()),
        "atualizadoEm": datetime.now().isoformat() + "Z",
    }
    if not kv_put_json("precos:meta", meta, ttl_days=8):
        erros += 1

    # 3b. Médias por município (único JSON com todos os municípios)
    payload_mun = {"s": semana_coleta, "m": municipios_media}
    if not kv_put_json("precos:municipios", payload_mun, ttl_days=8):
        erros += 1

    # 3c. Postos por UF (uma chave por UF para não estourar 25MB)
    for uf, postos in sorted(postos_por_uf.items()):
        key = f"precos:postos:{uf}"
        payload = {"s": semana_coleta, "postos": postos}
        if not kv_put_json(key, payload, ttl_days=8):
            erros += 1

    # Resumo
    print(f"\n{'='*60}")
    if erros == 0:
        print(f"✅ Sync concluído com sucesso!")
        print(f"   Semana: {semana_ini} a {semana_fim}")
        print(f"   Postos: {len(por_cnpj):,} | Municípios: {len(municipios_media):,}")
        print(f"   Chaves KV: {2 + len(postos_por_uf)} gravadas")
    else:
        print(f"⚠️  Sync concluído com {erros} erro(s)")

    return 0 if erros == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
